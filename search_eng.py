import glob
import os
import re
from openpyxl.reader.excel import load_workbook

root_dir = r"C:\Users\Admin\Documents"
pattern = os.path.join(root_dir, "historia *.*")

search_term = input("Search Term: ")

try:
    pattern_re = re.compile(search_term, re.IGNORECASE)
except re.error as e:
    print("Incorrect regex: {e}")
    exit(1)

def search_xlsx(path, regex):
    try:
        wb = load_workbook(path, data_only=True)

        for sheet in wb.worksheets:
            for r,row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for c,cell in enumerate(row, start=1):

                    if cell is None:
                        continue
                    text = str(cell)

                    if regex.search(text):
                        print(
                            f"📊 {path} | arkusz '{sheet.title}' "
                            f"| komórka {r}:{c}: {text}"
                        )
    except Exception as e:
        print(f" Błąd XLSX ({path}): {e}")

files = glob.glob(pattern)

for file in files:
    print(f"Searching {file}")

if not files:
    print("No files found")
    exit()

for path in files:
    ext = os.path.splitext(path)[1]

    if ext == ".xlsx":
        search_xlsx(path, pattern_re)
    else:
        pass